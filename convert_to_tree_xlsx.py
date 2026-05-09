"""
把 qa_output 目录下每个 _qa.json, 转成独立的 Excel, 并按 source_rel_path
在 Excel_最终交付/ 下完美复刻原有的多级目录结构.

输入(默认):
    qa_output/
        **/*_qa.json          (每个 JSON 内部有 source_rel_path)

输出(默认):
    Excel_最终交付/
        <还原 source_rel_path 的父目录>/
            <原 md 文件名>_QA对.xlsx

Excel 单表头(7 列):
    分类 | question | answers | 支撑原文 | document | page | text/img/table

用法:
    python convert_to_tree_xlsx.py
    python convert_to_tree_xlsx.py --input qa_output --output Excel_最终交付
"""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADERS = [
    "分类", "question", "answers", "支撑原文", "document", "page", "text/img/table"
]
# 每列目标宽度(字符数, 基于 Excel 列宽单位, 大致与字符数对应)
COLUMN_WIDTHS = {
    "分类": 10,
    "question": 40,
    "answers": 60,
    "支撑原文": 60,
    "document": 30,
    "page": 8,
    "text/img/table": 14,
}
# 非法 Windows 文件名字符 -> 替换符
_WIN_INVALID = '<>:"/\\|?*'

# 从 evidence 里抠"第X页"样式的页码
_PAGE_PATTERNS = [
    re.compile(r"第\s*([0-9]{1,4})\s*[页頁]"),                # 第 12 页 / 第12頁
    re.compile(r"(?:page|p\.?)\s*([0-9]{1,4})", re.IGNORECASE),  # page 12 / p.12 / P 12
    re.compile(r"([0-9]{1,4})\s*[页頁]"),                      # 12页 (兜底, 保守放后面)
]


def extract_page_from_evidence(evidence: str) -> str:
    """从 evidence 里尝试提取页码, 提取不到返回 'N/A'"""
    if not evidence:
        return "N/A"
    for pat in _PAGE_PATTERNS:
        m = pat.search(evidence)
        if m:
            return m.group(1)
    return "N/A"


def sanitize_segment(seg: str) -> str:
    """清洗单段路径中的非法字符(用于 Windows 目录/文件名)"""
    out = []
    for ch in seg:
        if ch in _WIN_INVALID or ord(ch) < 32:
            out.append("_")
        else:
            out.append(ch)
    name = "".join(out).strip(" .") or "_"
    return name


def build_target_xlsx_path(
    source_rel_path: str, output_root: Path, fallback_stem: str = None
) -> Path:
    """
    source_rel_path 形如: 非标准文本文件/太平洋保險/世代悅享2產品手冊/世代悅享2產品手冊.md
    返回:  <output_root>/非标准文本文件/太平洋保險/世代悅享2產品手冊/世代悅享2產品手冊_QA对.xlsx
    """
    p = Path(source_rel_path)
    stem = p.stem or fallback_stem or "未知文件"
    parts = [sanitize_segment(s) for s in p.parent.parts]
    out_dir = output_root.joinpath(*parts) if parts else output_root
    return out_dir / f"{sanitize_segment(stem)}_QA对.xlsx"


def write_xlsx(qa_pairs: list, xlsx_path: Path):
    """写一个 xlsx, 6 列单表头 + 自动换行 + 首行加粗 + 合适列宽"""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "QA"

    # 1) 表头
    ws.append(HEADERS)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align

    # 2) 数据
    body_align = Alignment(vertical="top", wrap_text=True)
    for rec in qa_pairs:
        evidence = rec.get("evidence", "") or ""
        # page: 优先用 JSON 已有的非 N/A 值; 否则尝试从 evidence 抠; 都不行填 N/A
        raw_page = rec.get("page", "N/A")
        page_val = str(raw_page).strip() if raw_page is not None else ""
        if not page_val or page_val.upper() == "N/A":
            page_val = extract_page_from_evidence(evidence)

        row = [
            rec.get("category", ""),
            rec.get("question", ""),
            rec.get("answers", ""),
            evidence,
            rec.get("document", ""),
            page_val,
            rec.get("support_type", "text"),
        ]
        ws.append(row)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = body_align

    # 3) 列宽 + 冻结首行
    for col_idx, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[name]
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)


def convert_all(input_dir: Path, output_dir: Path) -> tuple:
    """扫描 input_dir 下所有 *_qa.json, 各自生成 xlsx. 返回 (文件数, QA 总数, 跳过数)"""
    if not input_dir.exists():
        print(f"[错误] 输入目录不存在: {input_dir}", file=sys.stderr)
        sys.exit(1)

    qa_files = sorted(input_dir.rglob("*_qa.json"))
    if not qa_files:
        print(f"[警告] {input_dir} 下没有 *_qa.json")
        return 0, 0, 0

    output_dir.mkdir(parents=True, exist_ok=True)

    n_files = 0
    n_qa = 0
    n_skip = 0
    for qa_file in qa_files:
        try:
            payload = json.loads(qa_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            print(f"[跳过] JSON 解析失败: {qa_file}: {e}")
            n_skip += 1
            continue

        qa_pairs = payload.get("qa_pairs", [])
        if not qa_pairs:
            print(f"[跳过] qa_pairs 为空: {qa_file}")
            n_skip += 1
            continue

        # source_rel_path 缺失时, 用 qa_file 相对 input_dir 的位置回退, 并去掉 _qa 后缀
        src_rel = payload.get("source_rel_path")
        if not src_rel:
            rel_from_input = qa_file.relative_to(input_dir)
            fallback_stem = rel_from_input.stem
            if fallback_stem.endswith("_qa"):
                fallback_stem = fallback_stem[:-3]
            src_rel = str(rel_from_input.with_name(fallback_stem + ".md"))

        xlsx_path = build_target_xlsx_path(src_rel, output_dir)
        write_xlsx(qa_pairs, xlsx_path)
        n_files += 1
        n_qa += len(qa_pairs)
        print(f"  [ok] {xlsx_path}  ({len(qa_pairs)} 条)")

    return n_files, n_qa, n_skip


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", default="qa_output", help="JSON 输入根目录")
    ap.add_argument(
        "--output", default="Excel_最终交付", help="Excel 输出根目录"
    )
    args = ap.parse_args()

    input_dir = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()

    print(f"输入: {input_dir}")
    print(f"输出: {output_dir}")
    print()

    n_files, n_qa, n_skip = convert_all(input_dir, output_dir)

    print()
    print("=== 完成 ===")
    print(f"写出 xlsx:   {n_files} 个")
    print(f"写出 QA 总数: {n_qa} 条")
    print(f"跳过:         {n_skip} 个")
    print(f"输出根目录:   {output_dir}")


if __name__ == "__main__":
    main()
